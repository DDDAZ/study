from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
import urllib.parse as up

headers = {'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.8',
            'Cache-Control': 'max-age=0',
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome'
                          '/48.0.2564.116 Safari/537.36',
            'Connection': 'keep-alive',
            'Referer': 'http://www.baidu.com/'
           }
headers2 = {
            "Proxy-Connection": "keep-alive",
            "Pragma": "no-cache",
            "Cache-Control": "no-cache",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome"
                          "/52.0.2743.116 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "DNT": "1",
            "Accept-Encoding": "gzip, deflate, sdch",
            "Accept-Language": "zh-CN,zh;q=0.8,en-US;q=0.6,en;q=0.4",
            "Referer": "https://www.baidu.com/s?wd=%BC%96%E7%A0%81&rsv_spt=1&rsv_iqid=0x9fcbc99a0000b5d7&issp=1"
                       "&f=8&rsv_bp=1&rsv_idx=2&ie=utf-8&rqlang=cn&tn=baiduhome_pg&rsv_enter=0&oq=If-None-Match"
                       "&inputT=7282&rsv_t=3001MlX2aUzape9perXDW%2FezcxiDTWU4Bt%2FciwbikdOLQHYY98rhPyD2LDNevDKyLLg2"
                       "&rsv_pq=c4163a510000b68a&rsv_sug3=24&rsv_sug1=14&rsv_sug7=100&rsv_sug2=0&rsv_sug4=7283",
            "Accept-Charset": "gb2312,gbk;q=0.7,utf-8;q=0.7,*;q=0.7",
            }
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome'
                     '/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


# 获取按类别的全部单页基础爬虫用链接，返回2维列表
def urls_get(lists):
    urls_list = []
    tags_list = lists
    for tag in tags_list:
        url_list = []
        url_list.append(tag)
        urls_tag = 'https://www.douban.com/tag/' + up.quote(tag) + '/book'
        for i in range(2):
            url = urls_tag + '?start=' + str(i*15)
            url_list.append(url)
        urls_list.append(url_list)
    return urls_list


# 接受2维列表作为参数，对其中每一类别的单页链接进行提取并处理爬取得到信息，调用存储函数进行保存
def deal(urls):
    count_sheet = 0
    wb = Workbook()
    for i in range(len(urls)):
        count = 0
        books_list = []
        for url in urls[i]:
            if count == 0:
                count += 1  # 跳过第一个tag标签：书籍类别
                continue
            else:
                r = requests.get(url, headers=headers2)
                r_bs = BeautifulSoup(r.text, features='lxml')
                html_tag = re.compile(r'</?[^>]+>')
                attr = r_bs.find_all('dl')

                # 以一本书为独立单元进行属性列表的构建，对传入的链接参数(单页)进行for循环遍历即可获得单页上的所有书信息列表
                for one in attr:
                    one_list = []

                    # 书名----0
                    title_ini = one.find('a', 'title')
                    title_str = str(title_ini)
                    title_sub = html_tag.sub('', title_str)
                    one_list.append(title_sub)

                    # 评分----1
                    rating_ini = one.find('span', 'rating_nums')
                    rating_str = str(rating_ini)
                    rating_sub = html_tag.sub('', rating_str)
                    one_list.append(rating_sub)

                    # 价格，出版日期，出版社，作者----2 3 4 5
                    four_ini = one.find('div', 'desc')
                    four_str = str(four_ini)
                    four_sub = html_tag.sub('', four_str)
                    four_strip = four_sub.strip(' \n')
                    four_split = re.split(' / ', four_strip)
                    one_list.append(four_split[-1])
                    one_list.append(four_split[-2])
                    one_list.append(four_split[-3])
                    one_list.append(four_split[0])

                    # 评论人数----6
                    href_ini = one.find('a').get('href')
                    r_href = requests.get(href_ini, headers=headers2)
                    r_href_bs = BeautifulSoup(r_href.text, features='lxml')
                    critic_ini = r_href_bs.find('span', {'property': 'v:votes'})
                    critic_str = str(critic_ini)
                    critic_sub = html_tag.sub('', critic_str)
                    one_list.append(critic_sub)

                    # 添加到总列表
                    books_list.append(one_list)
                    count += 1

        # 调用内部存储函数对爬取的单标签结果list进行存储到excel
        if count_sheet == 0:
            ws = wb.create_sheet(urls[i][0])
            for j in range(len(books_list)):
                ws.append(books_list[j])
            wb.save('books.xlsx')
        else:
            wb = load_workbook('books.xlsx')
            wb.create_sheet(urls[i][0]).append(books_list)
            ws = wb.create_sheet(urls[i][0])
            for j in range(len(books_list)):
                ws.append(books_list[j])
            wb.save('books.xlsx')
    wb.close()


if __name__ == '__main__':
    tags_str = input('请输入需要下载的图书类别（用，隔开）：')
    tags_list = tags_str.split('，')
    urls = urls_get(tags_list)
    deal(urls)
